"""
Controle de Pacientes - App Web Local
Desenvolvido para psicólogos gerenciarem seus atendimentos.
v1.2.2 - Correção de porta e abertura do navegador
"""

import io
import csv
import os
import re
import shutil
import sys
import logging
from logging.handlers import RotatingFileHandler
from datetime import datetime, date
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, Response, send_file
from flask_sqlalchemy import SQLAlchemy

# Determine data directory (user's home)
if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

DATA_DIR = os.path.join(BASE_DIR, "dados")
os.makedirs(DATA_DIR, exist_ok=True)

LOG_DIR = os.path.join(DATA_DIR, "logs")
os.makedirs(LOG_DIR, exist_ok=True)

app = Flask(__name__)
app.config['SECRET_KEY'] = 'psico-controle-local-2024'
app.config['SQLALCHEMY_DATABASE_URI'] = f"sqlite:///{os.path.join(DATA_DIR, 'pacientes.db')}"
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)


# ==================== LOGGING ====================
# Log em arquivo com rotação automática (para análise e troubleshooting)

log_handler = RotatingFileHandler(
    os.path.join(LOG_DIR, 'app.log'), maxBytes=1_000_000, backupCount=5, encoding='utf-8'
)
log_handler.setFormatter(logging.Formatter('%(asctime)s [%(levelname)s] %(message)s'))

logger = logging.getLogger('controle_pacientes')
logger.setLevel(logging.INFO)
logger.addHandler(log_handler)

app.logger.addHandler(log_handler)
app.logger.setLevel(logging.INFO)


# ==================== VALIDAÇÃO DE URL (Google Drive) ====================

DRIVE_URL_PATTERN = re.compile(r'^https://(drive|docs)\.google\.com/', re.IGNORECASE)


def _validar_url_drive(url):
    """Aceita vazio (campo opcional) ou uma URL do drive.google.com / docs.google.com."""
    if not url:
        return True
    return bool(DRIVE_URL_PATTERN.match(url.strip()))


# ==================== MODELS ====================

class Paciente(db.Model):
    __tablename__ = 'pacientes'

    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(200), nullable=False)
    data_nascimento = db.Column(db.Date, nullable=True)
    telefone = db.Column(db.String(20), nullable=True)
    email = db.Column(db.String(200), nullable=True)
    contato_emergencia = db.Column(db.String(200), nullable=True)
    tel_emergencia = db.Column(db.String(20), nullable=True)
    data_inicio = db.Column(db.Date, nullable=True)
    modalidade = db.Column(db.String(20), default='Presencial')  # Presencial, Online, Híbrido
    frequencia = db.Column(db.String(20), default='Semanal')  # Semanal, Quinzenal, Mensal
    dia_horario = db.Column(db.String(50), nullable=True)
    valor_sessao = db.Column(db.Float, default=0.0)
    forma_pagamento = db.Column(db.String(20), default='Particular')  # Particular, Convênio, PsyMeet
    status = db.Column(db.String(20), default='Ativo')  # Ativo, Pausado, Alta, Desistência, Encaminhado
    observacoes = db.Column(db.Text, nullable=True)
    criado_em = db.Column(db.DateTime, default=datetime.now)

    sessoes = db.relationship('Sessao', backref='paciente', lazy=True, cascade='all, delete-orphan',
                              order_by='Sessao.data.desc()')


class Sessao(db.Model):
    __tablename__ = 'sessoes'

    id = db.Column(db.Integer, primary_key=True)
    paciente_id = db.Column(db.Integer, db.ForeignKey('pacientes.id'), nullable=False)
    data = db.Column(db.Date, nullable=False)
    horario_inicio = db.Column(db.String(5), nullable=True)
    horario_fim = db.Column(db.String(5), nullable=True)
    presenca = db.Column(db.String(20), default='Compareceu')  # Compareceu, Faltou, Cancelou, Remarcou
    numero_sessao = db.Column(db.Integer, nullable=True)
    pago = db.Column(db.Boolean, default=False)
    valor_pago = db.Column(db.Float, default=0.0)
    forma_pagamento = db.Column(db.String(30), nullable=True)
    evolucao = db.Column(db.Text, nullable=True)  # v1.2.0: evolução do paciente na sessão
    observacoes = db.Column(db.Text, nullable=True)
    link_prontuario = db.Column(db.Text, nullable=True)  # v1.2.0: link do Prontuário da Sessão (Google Drive)
    link_evolucao = db.Column(db.Text, nullable=True)  # v1.2.0: link da Evolução do Paciente (Google Drive)


# ==================== ROUTES ====================

@app.route('/')
def index():
    """Dashboard principal"""
    total_pacientes = Paciente.query.filter_by(status='Ativo').count()
    total_sessoes_mes = Sessao.query.filter(
        db.extract('month', Sessao.data) == date.today().month,
        db.extract('year', Sessao.data) == date.today().year
    ).count()
    receita_mes = db.session.query(db.func.sum(Sessao.valor_pago)).filter(
        db.extract('month', Sessao.data) == date.today().month,
        db.extract('year', Sessao.data) == date.today().year,
        Sessao.pago == True
    ).scalar() or 0.0
    a_receber = db.session.query(db.func.sum(Sessao.valor_pago)).filter(
        Sessao.pago == False,
        Sessao.presenca == 'Compareceu'
    ).scalar() or 0.0

    proximas_sessoes = Sessao.query.filter(
        Sessao.data >= date.today()
    ).order_by(Sessao.data.asc()).limit(5).all()

    return render_template('index.html',
                           total_pacientes=total_pacientes,
                           total_sessoes_mes=total_sessoes_mes,
                           receita_mes=receita_mes,
                           a_receber=a_receber,
                           proximas_sessoes=proximas_sessoes)


# --- Pacientes ---

@app.route('/pacientes')
def listar_pacientes():
    status_filter = request.args.get('status', 'Ativo')
    busca = request.args.get('busca', '').strip()

    query = Paciente.query
    if status_filter != 'Todos':
        query = query.filter_by(status=status_filter)
    if busca:
        query = query.filter(Paciente.nome.ilike(f'%{busca}%'))

    pacientes = query.order_by(Paciente.nome).all()
    return render_template('pacientes.html', pacientes=pacientes, status_filter=status_filter, busca=busca)


@app.route('/pacientes/novo', methods=['GET', 'POST'])
def novo_paciente():
    if request.method == 'POST':
        paciente = Paciente(
            nome=request.form['nome'],
            data_nascimento=_parse_date(request.form.get('data_nascimento')),
            telefone=request.form.get('telefone'),
            email=request.form.get('email'),
            contato_emergencia=request.form.get('contato_emergencia'),
            tel_emergencia=request.form.get('tel_emergencia'),
            data_inicio=_parse_date(request.form.get('data_inicio')),
            modalidade=request.form.get('modalidade', 'Presencial'),
            frequencia=request.form.get('frequencia', 'Semanal'),
            dia_horario=request.form.get('dia_horario'),
            valor_sessao=float(request.form.get('valor_sessao') or 0),
            forma_pagamento=request.form.get('forma_pagamento', 'Particular'),
            status='Ativo',
            observacoes=request.form.get('observacoes')
        )
        db.session.add(paciente)
        db.session.commit()
        logger.info(f"Paciente criado: id={paciente.id} nome={paciente.nome}")
        flash('Paciente cadastrado com sucesso!', 'success')
        return redirect(url_for('listar_pacientes'))
    return render_template('paciente_form.html', paciente=None)


@app.route('/pacientes/<int:id>/editar', methods=['GET', 'POST'])
def editar_paciente(id):
    paciente = Paciente.query.get_or_404(id)
    if request.method == 'POST':
        paciente.nome = request.form['nome']
        paciente.data_nascimento = _parse_date(request.form.get('data_nascimento'))
        paciente.telefone = request.form.get('telefone')
        paciente.email = request.form.get('email')
        paciente.contato_emergencia = request.form.get('contato_emergencia')
        paciente.tel_emergencia = request.form.get('tel_emergencia')
        paciente.data_inicio = _parse_date(request.form.get('data_inicio'))
        paciente.modalidade = request.form.get('modalidade', 'Presencial')
        paciente.frequencia = request.form.get('frequencia', 'Semanal')
        paciente.dia_horario = request.form.get('dia_horario')
        paciente.valor_sessao = float(request.form.get('valor_sessao') or 0)
        paciente.forma_pagamento = request.form.get('forma_pagamento', 'Particular')
        paciente.status = request.form.get('status', 'Ativo')
        paciente.observacoes = request.form.get('observacoes')
        db.session.commit()
        logger.info(f"Paciente atualizado: id={paciente.id} nome={paciente.nome}")
        flash('Paciente atualizado!', 'success')
        return redirect(url_for('listar_pacientes'))
    return render_template('paciente_form.html', paciente=paciente)


@app.route('/pacientes/<int:id>')
def ver_paciente(id):
    paciente = Paciente.query.get_or_404(id)
    sessoes = Sessao.query.filter_by(paciente_id=id).order_by(Sessao.data.desc()).all()
    total_sessoes = len([s for s in sessoes if s.presenca == 'Compareceu'])
    return render_template('paciente_detalhe.html', paciente=paciente, sessoes=sessoes, total_sessoes=total_sessoes)


@app.route('/pacientes/<int:id>/excluir', methods=['POST'])
def excluir_paciente(id):
    paciente = Paciente.query.get_or_404(id)
    logger.info(f"Paciente excluído: id={paciente.id} nome={paciente.nome}")
    db.session.delete(paciente)
    db.session.commit()
    flash(f'Paciente "{paciente.nome}" excluído.', 'success')
    return redirect(url_for('listar_pacientes'))


# --- Sessões ---

@app.route('/sessoes')
def listar_sessoes():
    presenca_filter = request.args.get('presenca', 'Todas')
    paciente_id_filter = request.args.get('paciente_id', '').strip()
    pago_filter = request.args.get('pago', '').strip()
    data_inicio_str = request.args.get('data_inicio', '').strip()
    data_fim_str = request.args.get('data_fim', '').strip()

    query = Sessao.query.join(Paciente)
    if presenca_filter != 'Todas':
        query = query.filter(Sessao.presenca == presenca_filter)
    if paciente_id_filter:
        query = query.filter(Sessao.paciente_id == int(paciente_id_filter))
    if pago_filter in ('Sim', 'Não'):
        query = query.filter(Sessao.pago == (pago_filter == 'Sim'))
    data_inicio = _parse_date(data_inicio_str)
    if data_inicio:
        query = query.filter(Sessao.data >= data_inicio)
    data_fim = _parse_date(data_fim_str)
    if data_fim:
        query = query.filter(Sessao.data <= data_fim)

    sessoes = query.order_by(Sessao.data.desc()).all()
    pacientes = Paciente.query.order_by(Paciente.nome).all()
    return render_template('sessoes.html',
                           sessoes=sessoes,
                           pacientes=pacientes,
                           presenca_filter=presenca_filter,
                           paciente_id_filter=paciente_id_filter,
                           pago_filter=pago_filter,
                           data_inicio=data_inicio_str,
                           data_fim=data_fim_str)


@app.route('/sessoes/nova', methods=['GET', 'POST'])
def nova_sessao():
    if request.method == 'POST':
        paciente_id = int(request.form['paciente_id'])
        data = _parse_date(request.form.get('data'))
        if not data:
            flash('A data da sessão é obrigatória.', 'error')
            return redirect(request.url)
        link_prontuario = request.form.get('link_prontuario', '').strip() or None
        link_evolucao = request.form.get('link_evolucao', '').strip() or None
        if not _validar_url_drive(link_prontuario) or not _validar_url_drive(link_evolucao):
            flash('Os links de Prontuário/Evolução devem ser URLs do Google Drive (drive.google.com ou docs.google.com).', 'error')
            return redirect(request.url)
        # Calculate session number
        count = Sessao.query.filter_by(paciente_id=paciente_id, presenca='Compareceu').count()
        sessao = Sessao(
            paciente_id=paciente_id,
            data=data,
            horario_inicio=request.form.get('horario_inicio'),
            horario_fim=request.form.get('horario_fim'),
            presenca=request.form.get('presenca', 'Compareceu'),
            numero_sessao=count + 1 if request.form.get('presenca') == 'Compareceu' else None,
            pago=request.form.get('pago') == 'Sim',
            valor_pago=float(request.form.get('valor_pago') or 0),
            forma_pagamento=request.form.get('forma_pagamento'),
            evolucao=request.form.get('evolucao'),
            observacoes=request.form.get('observacoes'),
            link_prontuario=link_prontuario,
            link_evolucao=link_evolucao
        )
        db.session.add(sessao)
        db.session.commit()
        logger.info(f"Sessão criada: id={sessao.id} paciente_id={paciente_id} data={data}")
        flash('Sessão registrada!', 'success')
        return redirect(url_for('ver_paciente', id=paciente_id))
    pacientes = Paciente.query.filter_by(status='Ativo').order_by(Paciente.nome).all()
    paciente_id = request.args.get('paciente_id')
    return render_template('sessao_form.html', pacientes=pacientes, paciente_id=paciente_id)


@app.route('/sessoes/<int:id>/editar', methods=['GET', 'POST'])
def editar_sessao(id):
    sessao = Sessao.query.get_or_404(id)
    if request.method == 'POST':
        data = _parse_date(request.form.get('data'))
        if not data:
            flash('A data da sessão é obrigatória.', 'error')
            return redirect(request.url)
        link_prontuario = request.form.get('link_prontuario', '').strip() or None
        link_evolucao = request.form.get('link_evolucao', '').strip() or None
        if not _validar_url_drive(link_prontuario) or not _validar_url_drive(link_evolucao):
            flash('Os links de Prontuário/Evolução devem ser URLs do Google Drive (drive.google.com ou docs.google.com).', 'error')
            return redirect(request.url)
        sessao.data = data
        sessao.horario_inicio = request.form.get('horario_inicio')
        sessao.horario_fim = request.form.get('horario_fim')
        sessao.presenca = request.form.get('presenca', 'Compareceu')
        sessao.pago = request.form.get('pago') == 'Sim'
        sessao.valor_pago = float(request.form.get('valor_pago') or 0)
        sessao.forma_pagamento = request.form.get('forma_pagamento')
        sessao.evolucao = request.form.get('evolucao')
        sessao.observacoes = request.form.get('observacoes')
        sessao.link_prontuario = link_prontuario
        sessao.link_evolucao = link_evolucao
        db.session.commit()
        logger.info(f"Sessão atualizada: id={sessao.id} paciente_id={sessao.paciente_id}")
        flash('Sessão atualizada!', 'success')
        return redirect(url_for('ver_paciente', id=sessao.paciente_id))
    pacientes = Paciente.query.filter_by(status='Ativo').order_by(Paciente.nome).all()
    return render_template('sessao_form.html', sessao=sessao, pacientes=pacientes, paciente_id=sessao.paciente_id)


@app.route('/sessoes/<int:id>/excluir', methods=['POST'])
def excluir_sessao(id):
    sessao = Sessao.query.get_or_404(id)
    paciente_id = sessao.paciente_id
    logger.info(f"Sessão excluída: id={sessao.id} paciente_id={paciente_id}")
    db.session.delete(sessao)
    db.session.commit()
    flash('Sessão excluída.', 'success')
    return redirect(url_for('ver_paciente', id=paciente_id))


# --- Financeiro ---

@app.route('/financeiro')
def financeiro():
    # Monthly summary for current year
    ano = date.today().year
    meses = []
    for mes in range(1, 13):
        total = Sessao.query.filter(
            db.extract('month', Sessao.data) == mes,
            db.extract('year', Sessao.data) == ano
        ).count()
        receita = db.session.query(db.func.sum(Sessao.valor_pago)).filter(
            db.extract('month', Sessao.data) == mes,
            db.extract('year', Sessao.data) == ano,
            Sessao.pago == True
        ).scalar() or 0.0
        pendente = db.session.query(db.func.sum(Sessao.valor_pago)).filter(
            db.extract('month', Sessao.data) == mes,
            db.extract('year', Sessao.data) == ano,
            Sessao.pago == False,
            Sessao.presenca == 'Compareceu'
        ).scalar() or 0.0
        faltas = Sessao.query.filter(
            db.extract('month', Sessao.data) == mes,
            db.extract('year', Sessao.data) == ano,
            Sessao.presenca == 'Faltou'
        ).count()
        if total > 0:
            meses.append({
                'mes': mes,
                'nome': _nome_mes(mes),
                'total_sessoes': total,
                'receita': receita,
                'pendente': pendente,
                'faltas': faltas
            })
    return render_template('financeiro.html', meses=meses, ano=ano)


# --- Exportação ---

@app.route('/exportar')
def exportar():
    return render_template('exportar.html')


@app.route('/exportar/pacientes')
def exportar_pacientes():
    formato = request.args.get('formato', 'csv')
    pacientes = Paciente.query.order_by(Paciente.nome).all()

    headers = ['Nome', 'Data Nascimento', 'Telefone', 'E-mail', 'Contato Emergência',
               'Tel. Emergência', 'Data Início', 'Modalidade', 'Frequência', 'Dia/Horário',
               'Valor Sessão', 'Forma Pagamento', 'Status', 'Observações']

    rows = []
    for p in pacientes:
        rows.append([
            p.nome,
            p.data_nascimento.strftime('%d/%m/%Y') if p.data_nascimento else '',
            p.telefone or '',
            p.email or '',
            p.contato_emergencia or '',
            p.tel_emergencia or '',
            p.data_inicio.strftime('%d/%m/%Y') if p.data_inicio else '',
            p.modalidade,
            p.frequencia,
            p.dia_horario or '',
            f'{p.valor_sessao:.2f}',
            p.forma_pagamento,
            p.status,
            p.observacoes or ''
        ])

    if formato == 'excel':
        return _export_excel('pacientes', headers, rows)
    return _export_csv('pacientes', headers, rows)


@app.route('/exportar/sessoes')
def exportar_sessoes():
    formato = request.args.get('formato', 'csv')
    sessoes = Sessao.query.order_by(Sessao.data.desc()).all()

    headers = ['Paciente', 'Data', 'Horário Início', 'Horário Fim', 'Presença',
               'Nº Sessão', 'Pago', 'Valor Pago', 'Forma Pagamento', 'Evolução', 'Observações']

    rows = []
    for s in sessoes:
        rows.append([
            s.paciente.nome,
            s.data.strftime('%d/%m/%Y'),
            s.horario_inicio or '',
            s.horario_fim or '',
            s.presenca,
            str(s.numero_sessao) if s.numero_sessao else '',
            'Sim' if s.pago else 'Não',
            f'{s.valor_pago:.2f}',
            s.forma_pagamento or '',
            s.evolucao or '',
            s.observacoes or ''
        ])

    if formato == 'excel':
        return _export_excel('sessoes', headers, rows)
    return _export_csv('sessoes', headers, rows)


@app.route('/backup')
def backup():
    return render_template('backup.html')


@app.route('/backup/download')
def backup_download():
    db.session.remove()
    db.engine.dispose()
    db_path = os.path.join(DATA_DIR, 'pacientes.db')
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    return send_file(
        db_path,
        as_attachment=True,
        download_name=f'backup_pacientes_{timestamp}.db'
    )


@app.route('/backup/restaurar', methods=['POST'])
def backup_restaurar():
    arquivo = request.files.get('arquivo')
    if not arquivo or arquivo.filename == '':
        flash('Selecione um arquivo de backup (.db) para restaurar.', 'error')
        return redirect(url_for('backup'))

    if not arquivo.filename.lower().endswith('.db'):
        flash('Arquivo inválido. Selecione um arquivo .db de backup.', 'error')
        return redirect(url_for('backup'))

    conteudo = arquivo.read()
    if not conteudo.startswith(b'SQLite format 3\x00'):
        flash('Arquivo inválido: não é um banco de dados SQLite válido.', 'error')
        return redirect(url_for('backup'))

    db_path = os.path.join(DATA_DIR, 'pacientes.db')

    db.session.remove()
    db.engine.dispose()

    backups_dir = os.path.join(DATA_DIR, 'backups')
    os.makedirs(backups_dir, exist_ok=True)
    if os.path.exists(db_path):
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        shutil.copy2(db_path, os.path.join(backups_dir, f'pre_restauracao_{timestamp}.db'))

    with open(db_path, 'wb') as f:
        f.write(conteudo)

    logger.info("Banco de dados restaurado a partir de backup enviado pelo usuário.")
    flash('Backup restaurado com sucesso! Uma cópia do banco anterior foi guardada em dados/backups.', 'success')
    return redirect(url_for('backup'))


# ==================== HELPERS ====================

def _parse_date(date_str):
    if not date_str:
        return None
    # Try DD/MM/YYYY (Brazilian format)
    try:
        return datetime.strptime(date_str.strip(), '%d/%m/%Y').date()
    except ValueError:
        pass
    # Try YYYY-MM-DD (ISO format, fallback)
    try:
        return datetime.strptime(date_str.strip(), '%Y-%m-%d').date()
    except ValueError:
        return None


def _nome_mes(mes):
    nomes = ['', 'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
             'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
    return nomes[mes]


def _export_csv(name, headers, rows):
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';', quoting=csv.QUOTE_MINIMAL)
    writer.writerow(headers)
    writer.writerows(rows)
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment; filename={name}_{date.today().strftime("%Y%m%d")}.csv'}
    )


def _export_excel(name, headers, rows):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = name.capitalize()

    # Header styling
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='6C63FF', end_color='6C63FF', fill_type='solid')

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-width columns
    for col in ws.columns:
        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'{name}_{date.today().strftime("%Y%m%d")}.xlsx'
    )


# ==================== MIGRAÇÕES ====================

def _run_migrations():
    """Adiciona colunas novas a tabelas já existentes, sem apagar dados.

    db.create_all() só cria tabelas ausentes — não altera tabelas que já
    existem. Bancos de produção (com pacientes/sessões já cadastrados)
    precisam desse passo para receber colunas adicionadas em versões novas.
    """
    inspector = db.inspect(db.engine)
    if 'sessoes' not in inspector.get_table_names():
        return  # tabela ainda não existe, será criada pelo create_all()

    colunas_existentes = {c['name'] for c in inspector.get_columns('sessoes')}
    colunas_novas = {
        'link_prontuario': 'TEXT',
        'link_evolucao': 'TEXT',
        'evolucao': 'TEXT',
    }
    with db.engine.connect() as conn:
        for nome, tipo in colunas_novas.items():
            if nome not in colunas_existentes:
                conn.execute(db.text(f'ALTER TABLE sessoes ADD COLUMN {nome} {tipo}'))
                conn.commit()
                logger.info(f"Migração: coluna '{nome}' adicionada à tabela 'sessoes'.")


# ==================== MAIN ====================

if __name__ == '__main__':
    with app.app_context():
        _run_migrations()
        db.create_all()
    logger.info("Aplicação iniciada - v1.2.2")
    print("\n  Controle de Pacientes v1.2.2")
    print("  Acesse no navegador: http://127.0.0.1:8642\n")
    app.run(host='127.0.0.1', port=8642, debug=False)
